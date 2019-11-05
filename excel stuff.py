import openpyxl, os
from openpyxl.styles import Font, Border, Side, PatternFill


os.chdir(r'C:\Github local repos\Survey_Admin_create\private')

wb = openpyxl.Workbook()

sheet1 = wb.active
sheet1.title = f'Redirects'
sheet1['A1'] = f'Redirects for '
sheet1['B2'] = 'Complete:'
sheet1['B3'] = 'Screened:'
sheet1['B4'] = 'Quota Full:'
sheet1['C2'] = 'completes redirect goes here'
sheet1['C3'] = 'SO redirect goes here'
sheet1['C4'] = 'QF redirect goes here'
sheet1.column_dimensions['B'].width = 15
sheet1.column_dimensions['C'].width = 95

emboldened = Font(bold=True)
sheet1['A1'].font = emboldened
sheet1['B2'].font = emboldened
sheet1['B3'].font = emboldened
sheet1['B4'].font = emboldened

thin = Side(border_style='thin')
surrounded = Border(top=thin, left=thin, right=thin, bottom=thin)
sheet1['B2'].border = surrounded
sheet1['B3'].border = surrounded
sheet1['B4'].border = surrounded
sheet1['C2'].border = surrounded
sheet1['C3'].border = surrounded
sheet1['C4'].border = surrounded

green = PatternFill("solid", fgColor="F0FFF0")
sheet1['B2'].fill = green
sheet1['C2'].fill = green

orange = PatternFill("solid", fgColor="FFF0F5")
sheet1['B3'].fill = orange
sheet1['C3'].fill = orange

red = PatternFill("solid", fgColor="FFEFD5")
sheet1['B4'].fill = red
sheet1['C4'].fill = red

wb.save('test.xlsx')