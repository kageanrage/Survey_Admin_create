import os, time, pprint, logging, sqlite3, subprocess, pyautogui, send2trash, datetime, calendar, sys, zcrmsdk, pyperclip, shutil
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from config import Config   # this imports the config file where the private data sits
import pandas as pd
from dateutil.relativedelta import *
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from zcrmsdk import *
from pprint import pprint
import se_general, se_admin, se_zoho
import cmv_quotas
from selenium.webdriver.common.by import By


# to avoid errors:
# Client Name in xls must be populated, and that column must be empty in excel from that new job onwards

logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')  # turns on logging
# logging.disable(logging.CRITICAL)     # switches off logging when desired

cfg = Config()  # create an instance of the Config class, essentially brings private config data into play
os.chdir(cfg.cwd)  # change the current working directory to the one stipulated in config file


# FUNCTIONS ################################################

def generate_dates_sa():
    now = datetime.datetime.now()  # current date and time as datetime object
    start_date_string = now.strftime("%d/%m/%Y %H:%M:%S")  # current date and time as string
    # print(f'Start Date is {start_date_string}')
    this_time_next_month = now + relativedelta(months=+1)  # date time object for today's date + time in a month
    next_month = this_time_next_month.month  # isolating just the number of next month
    year_next_month = this_time_next_month.year  # isolating just the year it will be next month
    year_next_month_plus_one = year_next_month + 1  # Added 02-11-21 to push end dates a year into future
    next_month_string = str(next_month)  # convert number of next month to a string
    if len(next_month_string) == 1:  # if next month is only 1 digit...
        next_month_string = "0" + next_month_string  # ...add a leading zero
    end_date_string = str(calendar.monthrange(year_next_month_plus_one, next_month)[1]) + "/" + str(next_month_string) \
                      + "/" + str(year_next_month_plus_one) + " 00:00:00"  # compile full end date string - changed 02-11-21 to push end dates a year into future
    this_time_last_month = now + relativedelta(months=-1)  # date time object for today's date + time a month ago
    last_month = this_time_last_month.month  # isolating just the number of last month
    year_last_month = this_time_last_month.year  # isolating just the year it was last month
    last_month_string = str(last_month)  # convert number of last month to a string
    if len(last_month_string) == 1:  # if last month is only 1 digit...
        last_month_string = "0" + last_month_string  # ...add a leading zero
    proposal_date_string = "01/" + str(last_month_string) \
                      + "/" + str(year_last_month)  # compile full prop date string
    return start_date_string, end_date_string, proposal_date_string


def generate_closing_date():
    print(f"close_date_raw is {close_date_raw}")
    close_date_trimmed = close_date_raw[0:10]
    print(f"close_date_trimmed is {close_date_trimmed}")
    close_date = datetime.datetime.strptime(close_date_trimmed, '%Y-%m-%d')
    last_day_in_close_month = calendar.monthrange(close_date.year, close_date.month)[1]
    close_month_string = str(close_date.month)  # convert number of close month to a string
    print(f"close_month_string is {close_month_string}")
    if len(close_month_string) == 1:  # if close month is only 1 digit...
        close_month_string = "0" + close_month_string  # ...add a leading zero
    print(f"After adding a leading zero where needed, close_month_string is now {close_month_string}")
    closing_date_string = str(last_day_in_close_month) + '/' + close_month_string + '/' + str(close_date.year)
    print(f"closing_date_string is {closing_date_string}")
    return closing_date_string


def date_reshuffler(original_date):
    logging.debug(f"original date is {original_date}")
    revised_date = original_date[6:10] + "-" + original_date[3:5] + "-" + original_date[0:2]
    logging.debug(f"revised date is {revised_date}")
    return revised_date


def search_contact_by_name(keyword):
    logging.debug('Function: search_contact_by_name')
    try:
        module_ins = ZCRMModule.get_instance('Contacts')  # module API Name
        resp = module_ins.search_records(keyword)  # search key word
        # print(resp.status_code)
        resp_info = resp.info
        # print('resp_info looks like this:')
        # print(resp_info)
        record_ins_arr = resp.data
        # print('record_ins_array looks like this:')
        # print(record_ins_arr)
        logging.debug(f'Number of contacts found with this name (i.e. len of record_ins_arr) is {len(record_ins_arr)}')
        assert len(record_ins_arr) == 1, 'Number of contacts found was not equal to 1'
        first_record = record_ins_arr[0]
        contact_id = first_record.entity_id

        record_ins_item_1_data = record_ins_arr[0].field_data
        # print('record looks like this:')
        # pprint(record_ins_item_1_data)
        full_name = record_ins_item_1_data['Full_Name']

        logging.debug(f"Returning Contact ID corresponding to {full_name}")
        logging.debug(f"ID found: {contact_id}")
        return contact_id

    except zcrmsdk.ZCRMException as ex:
        print(ex.status_code)
        print(ex.error_message)
        print(ex.error_code)
        print(ex.error_details)
        print(ex.error_content)


def create_potential():
    logging.debug('Function: create_potential')
    try:
        record = ZCRMRecord.get_instance('Potentials')  # module API Name

        record.set_field_value('Account_Name', client_name)
        record.set_field_value('Deal_Name', survey_name)
        record.set_field_value('Industry', industry)
        record.set_field_value('Account_Type', account_type)
        record.set_field_value('Campaign_Start_Date', campaign_start_date_api)  # note this format is different from the one I generated for the selenium input
        record.set_field_value('Campaign_End_Date', campaign_end_date_api)  # note this format is different from the one I generated for the selenium input
        record.set_field_value('Proposal_Date_Sent', proposal_date_api)  # note this format is different from the one I generated for the selenium input
        record.set_field_value('Closing_Date', closing_date_api)  # note this format is different from the one I generated for the selenium input
        record.set_field_value('Stage', stage)

        contact_id = search_contact_by_name(sales_contact)  # uses my fn to grab the contact's ID
        ac_dynamic_dict = {'name': sales_contact, 'id': contact_id}  # ... then pipes the ID and name into this dict for insertion into potential
        # print('ac_dynamic_dict looks like this:')
        # print(ac_dynamic_dict)
        # print('Attempting to set Contact_Name using dynamic dict')
        record.set_field_value('Contact_Name', ac_dynamic_dict)

        resp = record.create()
        new_potential_id = record.entity_id  # grabs the entity ID of the record, to then use this to look up the newly created potential
        print(resp.status_code)
        logging.debug(f"New potential's ID is {new_potential_id}")
        return new_potential_id

    except ZCRMException as ex:
        print(ex.status_code)
        print(ex.error_message)
        print(ex.error_code)
        print(ex.error_details)
        print(ex.error_content)


def get_potential_record_by_id(id):
    try:
        record = ZCRMRecord.get_instance('Potentials', id)
        resp = record.get()
        # print(resp.status_code)
        # print(f"entity ID for this record is {resp.data.entity_id}")
        # print(resp.data.created_by.id)
        # print(resp.data.modified_by.id)
        # print(resp.data.owner.id)
        # print(resp.data.created_by.name)
        # print(resp.data.created_time)
        # print(resp.data.modified_time)
        # print(resp.data.get_field_value('Email'))
        # print(resp.data.get_field_value('Last_Name'))
        # print('######### Attempting to display Full_Name - if this works I can return this value')
        deal_name = resp.data.get_field_value('Deal_Name')
        p_number = resp.data.get_field_value('Reference_Number')
        account_name_in_potential = resp.data.get_field_value('Account_Name')['name']
        logging.debug(f'account_name = {account_name_in_potential}')
        assert account_name_in_potential == client_name, "client names vary between potential and spreadsheet. Check for spaces etc?"
        # print(full_name)
        print('Data for the potential looks like this:')
        pprint(resp.data.field_data)
        # if resp.data.line_items is not None:
        #     for line_item in resp.data.line_items:
        #         print("::::::LINE ITEM DETAILS::::::")
        #         print(line_item.id)
        #         print(line_item.product.get_field_value('Product_Code'))
        return p_number

    except ZCRMException as ex:
        print(ex.status_code)
        print(ex.error_message)
        print(ex.error_code)
        print(ex.error_details)
        print(ex.error_content)





def create_redirects_xls(q, s, c):
    wb = openpyxl.Workbook()

    sheet1 = wb.active
    sheet1.title = f'Redirects - {p_number}'
    sheet1['A1'] = f'Redirects for {p_number} - {survey_name}'
    sheet1['B2'] = 'Complete:'
    sheet1['B3'] = 'Screened:'
    sheet1['B4'] = 'Quota Full:'
    sheet1['B5'] = 'Late Screened:'
    sheet1['B6'] = 'Quality Terminate:'
    sheet1['C2'] = str(cfg.redirect_prefix + c)
    sheet1['C3'] = str(cfg.redirect_prefix + s)
    sheet1['C4'] = str(cfg.redirect_prefix + q)
    sheet1['C5'] = str(cfg.redirect_prefix + ls)
    sheet1['C6'] = str(cfg.redirect_prefix + qt)
    sheet1.column_dimensions['B'].width = 15
    sheet1.column_dimensions['C'].width = 95

    emboldened = Font(bold=True)
    sheet1['A1'].font = emboldened
    sheet1['B2'].font = emboldened
    sheet1['B3'].font = emboldened
    sheet1['B4'].font = emboldened
    sheet1['B5'].font = emboldened
    sheet1['B6'].font = emboldened

    thin = Side(border_style='thin')
    surrounded = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet1['B2'].border = surrounded
    sheet1['B3'].border = surrounded
    sheet1['B4'].border = surrounded
    sheet1['B5'].border = surrounded
    sheet1['B6'].border = surrounded
    sheet1['C2'].border = surrounded
    sheet1['C3'].border = surrounded
    sheet1['C4'].border = surrounded
    sheet1['C5'].border = surrounded
    sheet1['C6'].border = surrounded

    green = PatternFill("solid", fgColor="F0FFF0")
    sheet1['B2'].fill = green
    sheet1['C2'].fill = green

    red = PatternFill("solid", fgColor="FFF0F5")
    sheet1['B3'].fill = red
    sheet1['C3'].fill = red

    orange = PatternFill("solid", fgColor="FFEFD5")
    sheet1['B4'].fill = orange
    sheet1['C4'].fill = orange

    # red = PatternFill("solid", fgColor="FFEFD5")
    sheet1['B5'].fill = red
    sheet1['C5'].fill = red

    # red = PatternFill("solid", fgColor="FFEFD5")
    sheet1['B6'].fill = red
    sheet1['C6'].fill = red



    wb.save(redirects_wb_path_name_ext)


def add_fields_to_redirects_xls():
    filename_inc_dir = f"{new_project_dir_path}\\{p_number} redirects.xlsx"  # changed this on 29-01-20
    wb = openpyxl.load_workbook(filename_inc_dir)  # changed this on 29-01-20

    sheet1 = wb.active

    sheet1['B9'] = 'P number:'
    sheet1['C9'] = p_number
    sheet1['B10'] = 'Survey ID:'
    sheet1['C10'] = survey_id

    emboldened = Font(bold=True)
    sheet1['B9'].font = emboldened
    sheet1['B10'].font = emboldened

    thin = Side(border_style='thin')
    surrounded = Border(top=thin, left=thin, right=thin, bottom=thin)
    sheet1['B9'].border = surrounded
    sheet1['B10'].border = surrounded
    sheet1['C9'].border = surrounded
    sheet1['C10'].border = surrounded

    wb.save(filename_inc_dir)


def enter_data_sa():
    driver.find_element('id', 'Name').send_keys(survey_name)  # using send_keys instead of script command here due to potential inclusion of apostrophes etc which stuff up the js syntax
    
    driver.find_element('id', 'Status').send_keys(str(status))
    driver.find_element('id', 'Title').send_keys(str(topic))
    driver.find_element('id', 'ProjectIONumber').send_keys(str(p_number))
    driver.find_element('id', 'ExpectedLength').send_keys(str(expected_loi))
    driver.find_element('id', 'ClientCompanyName').send_keys(str(client_name))
    driver.find_element('id', 'ExternalSurveyUrl').send_keys(str(external_survey_url))
    driver.find_element('id', 'StartDate').send_keys(str(start_date))
    driver.find_element('id', 'EndDate').send_keys(str(end_date))

    # Outcome messages
    driver.find_element('id', 'OutcomeFull').send_keys(str(qf_msg))
    driver.find_element('id', 'OutcomeScreened').send_keys(str(so_msg))
    driver.find_element('id', 'OutcomeComplete').send_keys(str(comp_msg))
    driver.find_element('id', 'OutcomeLateScreened').send_keys(str(so_msg))
    driver.find_element('id', 'OutcomeQualityTerminate').send_keys(str(so_msg))

    # Outcome value counts
    driver.find_element('id', 'OutcomeFullRewardValue').send_keys(str(prize_draw_entries))
    driver.find_element('id', 'OutcomeScreenedRewardValue').send_keys(str(prize_draw_entries))
    driver.find_element('id', 'OutcomeCompleteRewardValue').send_keys(str(prize_draw_entries))  # This will just input a '1'
    driver.find_element('id', 'OutcomeLateScreenedRewardValue').send_keys(str(prize_draw_entries))  
    driver.find_element('id', 'OutcomeQualityTerminateRewardValue').send_keys(str(prize_draw_entries))  

    # Outcome EC values
    driver.find_element('id', 'OutcomeCompleteSecondaryRewardValue').send_keys(str(edge_credits))
    driver.find_element('id', 'OutcomeFullSecondaryRewardValue').send_keys(str(ec_value_for_so_and_qf))
    driver.find_element('id', 'OutcomeScreenedSecondaryRewardValue').send_keys(str(ec_value_for_so_and_qf))
    driver.find_element('id', 'OutcomeLateScreenedSecondaryRewardValue').send_keys(str(ec_value_for_so_and_qf))
    driver.find_element('id', 'OutcomeQualityTerminateSecondaryRewardValue').send_keys(str(ec_value_for_so_and_qf))

    # Outcome Reward IDs
    driver.find_element('id', 'FullOutcomeRewardId').send_keys(qf_outcome_reward_id)  # this didn't work via JS execution method
    driver.find_element('id', 'ScreenedOutcomeRewardId').send_keys(so_outcome_reward_id)  
    driver.find_element('id', 'LateScreenedOutcomeRewardId').send_keys(so_outcome_reward_id)  
    driver.find_element('id', 'QualityTerminateOutcomeRewardId').send_keys(so_outcome_reward_id)  
    driver.find_element('id', 'CompleteOutcomeRewardId').send_keys(str_saying_comp_surv_reg_p_d)

    # I think these are now redundant
    # driver.find_element('id', 'OutcomeCompleteSecondaryRewardValue').value = '" + str(prize_draw_entries) + "';")
    # driver.find_element('id', 'OutcomeCompleteSecondaryRewardType').send_keys('Reward')
    # driver.find_element('id', 'OutcomeCompleteRewardType').send_keys(the_word_credits)

    if mobile_friendly == "No":
        driver.find_element('id', 'DesktopOnly').click()
    if len(str(survey_ids_to_exclude)) > 10:
        driver.find_element('id', 'ExcludePastSurveyIds').send_keys(survey_ids_to_exclude)  # amended 03-03-20
    # driver.find_element('id', 'TermsAndConditionsPdf').click()  # moved away from .click to ActionChains due to Chrome v78 bug
    tc_button = driver.find_element('id', 'TermsAndConditionsPdf')
    ActionChains(driver).click(tc_button).perform()
    time.sleep(2)
    pyautogui.typewrite(tc_filepath)  # since popup window is outside web browser, need a diff package to control
    time.sleep(2)
    pyautogui.press('enter')
    time.sleep(2)

    # add intro popup text
    tiny_mce = driver.find_element(By.CSS_SELECTOR, 'html')
    tiny_mce.click()
    time.sleep(1)
    pyautogui.typewrite(' ')
    time.sleep(1)
    pyautogui.typewrite(' ')
    time.sleep(1)
    ActionChains(driver).send_keys(Keys.ESCAPE).perform()

    time.sleep(3)
    if "CMV" in survey_name:
        # print(f"CMV is in survey_name: {survey_name}")
        pyautogui.typewrite(cfg.cmv_pre_survey_text)
        time.sleep(2)
    else:
        pyautogui.typewrite(cfg.default_pre_survey_text)
        time.sleep(2)

    submit_button = driver.find_element(By.CSS_SELECTOR, '#add-edit-survey > fieldset > dl > div.form_navigation > button')
    ActionChains(driver).click(submit_button).perform()
    # COMMENT OUT THE LAST ROW FOR TEST MODE, TO AVOID ACTUAL PROJECT CREATION  ###########


def grab_survey_id():
    project_sa_listing = driver.find_element(By.LINK_TEXT, survey_name)  # no longer works from Mar-20 with SA change to html table, but perhaps fixed 23-03    print('pausing for a few sec to make sure page has loaded')
    time.sleep(4)
    # use the following chunk if this method doesn't work and I need to sort the table and click first row:
    """
    # start_date_header = driver.find_element(By.CSS_SELECTOR, 'css=.col-06 > .mvc-grid-title')
    # start_date_header.click()
    # time.sleep(3)
    # start_date_header.click()
    # time.sleep(3)
    # project_sa_listing = driver.find_element(By.CSS_SELECTOR, ".selectable-row:nth-child(1) > .clickable:nth-child(1)")
    """
    project_sa_listing.click()
    time.sleep(4)
    current_url = driver.current_url
    s_id = current_url[59:]
    print(f's_id = {s_id}')
    assert len(s_id) == 36, f'The length of survey_id must be 36 and in fact is {len(s_id)}'
    return s_id


def create_test_quota():
    driver.get(f"https://data.studentedge.org/admin/survey/createquota?surveyId={survey_id}")
    time.sleep(2)
    name_field = driver.find_element('id', 'Name')
    name_field.send_keys('1')
    all_states_button = driver.find_element(By.LINK_TEXT, '>>')
    all_states_button.click()
    target_field = driver.find_element('id', 'Target')
    target_field.clear()
    target_field.send_keys('1')
    save_new_quota_button = driver.find_element(By.CSS_SELECTOR, '.green')
    save_new_quota_button.click()


def open_relevant_files():
    excel_name_path = cfg.live_excel_file_path + "\\" + cfg.live_excel_filename
    excel_file_name_path_ext = excel_name_path + ".xlsx"
    subprocess.Popen(f'explorer "{new_project_dir_path}"')  # opens new dir in windows explorer  # DISABLE FOR TESTING
    subprocess.Popen(f'explorer "{redirects_wb_path_name_ext}"')  # opens file in windows  # DISABLE FOR TESTING
    subprocess.Popen(f'explorer "{excel_file_name_path_ext}"')  # opens Survey Tracking file in windows, so I can add in project number manually  # DISABLE FOR TESTING


def determine_exclusion_survey_ids(survey_name):
    proj_dict = se_general.look_up_project(survey_name)
    if str(proj_dict['survey names to exclude']) != 'nan':  # if survey names listed for exclusion
        print('There are survey names listed in the xls, to exclude.')
        print('Survey names to exclude looks like this:')
        print(proj_dict['survey names to exclude'])
        str_of_survey_names_to_exclude = proj_dict['survey names to exclude']
        excl_surv_names = str_of_survey_names_to_exclude.split(',')

        survey_id_excl_list = []

        for s_name in excl_surv_names:
            # proj_dict for each of the survey ids I'm looking up
            proj_dict = se_general.look_up_project(s_name)
            survey_id_excl_list.append(proj_dict['Survey ID'])

        print('survey ids to exclude list looks like this:')
        print(survey_id_excl_list)
        survey_ids_to_exclude_str = ",".join(survey_id_excl_list)  # changed 31-03-20

    elif str(proj_dict['survey ids to exclude']) != 'nan':  # if survey_ids are manually listed for exclusion
        print('There are individual survey IDs listed in xls to exclude.')
        survey_id_excl_list = proj_dict['survey ids to exclude']
        print('survey ids to exclude list looks like this:')
        print(survey_id_excl_list)
        survey_ids_to_exclude_str = str(survey_id_excl_list)

    else:
        print('No survey ids to exclude')
        survey_ids_to_exclude_str = ""

    print('Survey exclusions string looks like this:')
    print(survey_ids_to_exclude_str)

    return survey_ids_to_exclude_str


proj_dict = se_general.look_up_latest_project()

# variables for Zoho + Survey Admin
survey_name = proj_dict['Survey Name']
topic = proj_dict['Topic']
expected_loi = str(int(proj_dict['Expected LOI']))
client_name = str(proj_dict['Client name'])
sales_contact = str(proj_dict['Sales Contact'])
edge_credits = str(int(proj_dict['Edge Credits']))
close_date_raw = str(proj_dict['Close month'])
mobile_friendly = str(proj_dict['mobile_friendly'])

survey_ids_to_exclude = determine_exclusion_survey_ids(survey_name)

# ZOHO variables
# Fixed variables - same for all projects
industry = "Other"
account_type = "Research Panel"
stage = "Closed Won - Signed IO Received"

# SURVEY ADMIN variables
# Fixed variables - same for all projects
status = 'Active'
qf_msg = cfg.qf_msg
so_msg = cfg.so_msg
comp_msg = cfg.comp_msg
external_survey_url = 'http://www.TO_BE_CONFIRMED.com'
prize_draw_entries = '1'
qf_outcome_reward_id = 'Disqualified Survey - Regular Prize Draw'
so_outcome_reward_id = 'Disqualified Survey - Regular Prize Draw'
str_saying_comp_surv_reg_p_d = 'Completed Survey - Regular Prize Draw'
the_word_credits = 'Credits'
tc_filepath = cfg.tc_filepath  # file path of T&Cs pdf file
ec_value_for_so_and_qf = "10"



# DATE VARIABLES
start_date, end_date, proposal_date = generate_dates_sa()  # generates start and end dates through the function
campaign_start_date = start_date[0:10]  # same as start date in survey admin, which is today's date, but trimmed
closing_date = generate_closing_date()
campaign_end_date = closing_date  # same as closing date
# format-adjusted date variables for use by API
proposal_date_api = date_reshuffler(proposal_date)
closing_date_api = date_reshuffler(closing_date)
campaign_start_date_api = date_reshuffler(campaign_start_date)
campaign_end_date_api = date_reshuffler(campaign_end_date)

se_general.check_for_bad_chars(survey_name, client_name, sales_contact)


# Zoho
oauth_client, refresh_token, user_identifier, oauth_tokens = se_zoho.init_zoho_api()

new_job_id = create_potential()  # create the new potential and store its ID in this variable
p_number = get_potential_record_by_id(new_job_id)  # use that ID to look up the newly created potential and store its P-number in this variable
# print(f"p-number for new project is {p_number}")

# Survey Admin variables
driver = se_general.init_selenium()
driver.implicitly_wait(30)

new_project_dir_path = cfg.projects_dir_path + "\\" + client_name + "\\" + p_number + " - " + survey_name
logging.debug('new_project_dir_path has now been created and it looks like this:')
logging.debug(new_project_dir_path)
redirects_wb_path_name_ext = new_project_dir_path + "\\" + p_number + " redirects.xlsx"

# Survey Admin + Windows levers
se_admin.login_sa_2fa(driver, cfg.create_survey_URL)  # now using fn from module

client_dir_path = cfg.projects_dir_path + "\\" + client_name
se_general.create_dir_if_not_exists(client_dir_path)
se_general.create_dir_if_not_exists(new_project_dir_path)

enter_data_sa()
survey_id = grab_survey_id()
qf, so, comp, ls, qt = se_admin.grab_redirects(driver, cfg.redirect_prefix, guid_only=False)
create_redirects_xls(qf, so, comp)
add_fields_to_redirects_xls()
create_test_quota()

# CMV quotas
if "CMV" in survey_name:
    print(f"CMV is in survey_name: {survey_name}")
    cmv_quotas.generate_cmv_quotas(cfg, driver)
    shutil.copy(cfg.cmv_quota_template, new_project_dir_path + "\\" + "cmv_quotas_template.xlsx")

driver.close()

open_relevant_files()
pyperclip.copy(f"{p_number} {survey_id}")  # copy p_number and survey_id to clipboard to then manually paste once script is done
